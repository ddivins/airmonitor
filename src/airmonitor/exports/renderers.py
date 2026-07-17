"""PNG, PDF, XLSX, CSV, and experiment-package rendering."""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from html import escape
from importlib.resources import as_file, files
import io
import json
import math
from pathlib import Path
import re
from tempfile import TemporaryDirectory
from typing import Any, Iterable
from zipfile import ZIP_DEFLATED, ZipFile

import matplotlib

matplotlib.use("Agg")
from matplotlib import dates as mdates  # noqa: E402
from matplotlib import image as mpimg  # noqa: E402
from matplotlib import pyplot as plt  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from reportlab.lib import colors  # noqa: E402
from reportlab.lib.enums import TA_CENTER  # noqa: E402
from reportlab.lib.pagesizes import letter, landscape  # noqa: E402
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # noqa: E402
from reportlab.lib.units import inch  # noqa: E402
from reportlab.platypus import (  # noqa: E402
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from airmonitor.exports.model import PrintExport


NAVY = "#061d49"
BLUE = "#1675d1"
TEAL = "#10b7b2"
INK = "#172845"
MUTED = "#66748a"
GRID = "#dbe4eb"
WARNING = "#8a5a00"
SAFE_EXPORT_COLUMNS = {
    "print": tuple(),
    "sgx": (
        "id", "sampled_at", "sensor_id", "session_id", "print_id",
        "sensor_protocol", "sensor_port", "gas_ppm", "gas_mass", "full_scale",
        "temperature_c", "humidity_rh",
    ),
    "sps30": (
        "id", "sampled_at", "sensor_id", "session_id", "sensor_port",
        "mass_pm1_0", "mass_pm2_5", "mass_pm4_0", "mass_pm10",
        "number_pm0_5", "number_pm1_0", "number_pm2_5", "number_pm4_0",
        "number_pm10", "typical_particle_size",
    ),
    "levoit": (
        "id", "sampled_at", "device_name", "power_state", "mode", "fan_level",
        "pm2_5", "air_quality", "filter_life_percent",
    ),
}


def safe_stem(report: PrintExport) -> str:
    name = re.sub(r"[^a-z0-9]+", "-", report.title.lower()).strip("-")
    name = (name[:60].rstrip("-") or "unnamed-print")
    date = report.started_at.strftime("%Y-%m-%d")
    return f"airmonitor-print-{report.print_id:04d}-{name}-{date}"


def render_publication_png(report: PrintExport, destination: str | Path) -> Path:
    destination = Path(destination)
    fig = plt.figure(figsize=(16, 9), dpi=240, facecolor="#f4f8fb")
    grid = fig.add_gridspec(
        4, 2, height_ratios=(0.78, 1.2, 1.2, 0.24),
        width_ratios=(3.25, 1.35), hspace=0.48, wspace=0.22,
        left=0.055, right=0.97, top=0.95, bottom=0.065,
    )

    header = fig.add_subplot(grid[0, :])
    header.axis("off")
    _draw_logo(header)
    header.text(0.18, 0.76, report.title, color=NAVY, fontsize=25, weight="bold", va="center")
    header.text(
        0.18, 0.43,
        f"Print #{report.print_id:04d}  |  {_format_range(report.started_at, report.ended_at)}"
        + ("  |  ACTIVE WHEN GENERATED" if report.active else ""),
        color=INK, fontsize=11.5, va="center",
    )
    header.text(
        0.18, 0.16,
        f"Export window: {_format_range(report.window_start, report.window_end)}  |  "
        f"{_material_line(report)}",
        color=MUTED, fontsize=10.2, va="center",
    )

    voc = fig.add_subplot(grid[1, 0])
    _plot_series(
        voc, report, report.sgx_samples,
        (("gas_ppm", "VOC", BLUE),),
        "VOC", "ppm",
    )
    pm = fig.add_subplot(grid[2, 0])
    _plot_series(
        pm, report, report.sps30_samples,
        (
            ("mass_pm1_0", "PM1", TEAL),
            ("mass_pm2_5", "PM2.5", BLUE),
            ("mass_pm4_0", "PM4", "#7754c5"),
            ("mass_pm10", "PM10", "#e49b00"),
        ),
        "Particulate matter mass", "ug/m3",
    )
    environment = fig.add_subplot(grid[1:3, 1])
    _plot_environment(environment, report)
    _summary_box(environment, report)

    footer = fig.add_subplot(grid[3, :])
    footer.axis("off")
    footer.text(
        0, 0.52,
        "AirMonitor VOC measurements are intended for relative/comparative analysis and are not "
        "compound-specific, regulatory, medical, OSHA-valid, or life-safety exposure measurements.",
        color=MUTED, fontsize=8.3, va="center",
    )
    footer.text(
        1, 0.52,
        f"Generated {report.generated_at.strftime('%Y-%m-%d %H:%M UTC')} | "
        f"AirMonitor {report.project_version} | {report.git_commit}",
        color=MUTED, fontsize=8.3, ha="right", va="center",
    )
    fig.savefig(destination, dpi=240, facecolor=fig.get_facecolor())
    plt.close(fig)
    return destination


def render_pdf(report: PrintExport, destination: str | Path) -> Path:
    destination = Path(destination)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="AirTitle", parent=styles["Title"], textColor=colors.HexColor(NAVY),
        fontSize=27, leading=31, spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        name="AirSub", parent=styles["BodyText"], textColor=colors.HexColor(MUTED),
        fontSize=10, leading=14,
    ))
    styles.add(ParagraphStyle(
        name="AirHeading", parent=styles["Heading2"], textColor=colors.HexColor(NAVY),
        fontSize=16, leading=20, spaceBefore=4, spaceAfter=9,
    ))
    styles.add(ParagraphStyle(
        name="AirFooter", parent=styles["BodyText"], textColor=colors.HexColor(MUTED),
        fontSize=8, leading=10, alignment=TA_CENTER,
    ))
    doc = SimpleDocTemplate(
        str(destination), pagesize=landscape(letter),
        rightMargin=0.55 * inch, leftMargin=0.55 * inch,
        topMargin=0.45 * inch, bottomMargin=0.45 * inch,
        title=f"AirMonitor Print {report.print_id}: {report.title}",
        author="AirMonitor",
    )
    story: list[Any] = []
    logo = files("airmonitor").joinpath("status_static/airmonitor-logo-300px.webp")
    with as_file(logo) as logo_path:
        story.append(Image(str(logo_path), width=2.0 * inch, height=0.98 * inch))
    story.extend([
        Paragraph(escape(report.title), styles["AirTitle"]),
        Paragraph(
            f"AirMonitor print experiment report - Print #{report.print_id:04d}"
            + (" - generated while active" if report.active else ""),
            styles["AirSub"],
        ),
        Spacer(1, 0.15 * inch),
        _pdf_table([
            ("Print start", _display_time(report.started_at)),
            ("Print end / last seen", _display_time(report.ended_at)),
            ("Export window", _format_range(report.window_start, report.window_end)),
            ("Material", _value(report.print_record.get("filament_type"))),
            ("Filament profile", _value(report.print_record.get("filament_profile") or report.print_record.get("filament_sub_brand"))),
            ("Filament color", _value(report.print_record.get("filament_color"))),
            ("Final state", _value(report.print_record.get("ended_gcode_state") or report.print_record.get("last_gcode_state"))),
        ], widths=(1.65 * inch, 7.2 * inch)),
        Spacer(1, 0.18 * inch),
        Paragraph("Summary statistics", styles["AirHeading"]),
        _metric_table(report),
        Spacer(1, 0.18 * inch),
        Paragraph(
            "Methodology: the report window begins 30 minutes before the print start and ends "
            "30 minutes after the recorded print end (or last-seen time for an active print). "
            "Baselines are medians of valid pre-print samples.",
            styles["AirSub"],
        ),
    ])
    if report.warnings:
        story.extend([
            Spacer(1, 0.12 * inch),
            Paragraph("Warnings: " + " ".join(escape(item) for item in report.warnings), styles["AirSub"]),
        ])

    with TemporaryDirectory(prefix="airmonitor-pdf-") as directory:
        charts = Path(directory) / "charts.png"
        _render_pdf_charts(report, charts)
        story.extend([
            PageBreak(),
            Paragraph("Sensor history", styles["AirHeading"]),
            Image(str(charts), width=9.7 * inch, height=6.45 * inch),
            PageBreak(),
            Paragraph("Interpretation and provenance", styles["AirHeading"]),
            Paragraph(
                "VOC values are cross-sensitive total-VOC measurements intended for relative and "
                "comparative analysis. They do not identify styrene or another individual compound "
                "and must not be compared directly with compound-specific OSHA or NIOSH limits.",
                styles["BodyText"],
            ),
            Spacer(1, 0.18 * inch),
            _pdf_table([
                ("AirMonitor version", report.project_version),
                ("Git commit", report.git_commit),
                ("Generated", _display_time(report.generated_at)),
                ("SGX samples", str(len(report.sgx_samples))),
                ("SPS30 samples", str(len(report.sps30_samples))),
                ("Levoit samples", str(len(report.levoit_samples))),
            ], widths=(1.65 * inch, 7.2 * inch)),
            Spacer(1, 0.25 * inch),
            Paragraph(
                "AirMonitor is a DIY monitoring platform and is not a certified regulatory, medical, "
                "occupational-exposure, or life-safety instrument.",
                styles["AirFooter"],
            ),
        ])
        doc.build(story, onFirstPage=_pdf_page, onLaterPages=_pdf_page)
    return destination


def render_xlsx(report: PrintExport, destination: str | Path) -> Path:
    destination = Path(destination)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _summary_sheet(summary, report)
    _raw_sheet(workbook.create_sheet("SGX Samples"), report.sgx_samples, SAFE_EXPORT_COLUMNS["sgx"])
    _raw_sheet(workbook.create_sheet("SPS30 Samples"), report.sps30_samples, SAFE_EXPORT_COLUMNS["sps30"])
    _metadata_sheet(workbook.create_sheet("Print Metadata"), report)
    if report.levoit_samples:
        _raw_sheet(workbook.create_sheet("Filter State"), report.levoit_samples, SAFE_EXPORT_COLUMNS["levoit"])
        workbook["Filter State"].insert_rows(1)
        workbook["Filter State"].merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SAFE_EXPORT_COLUMNS["levoit"]))
        workbook["Filter State"]["A1"] = "Historical Levoit samples only; Bento history is not currently recorded."
        workbook["Filter State"]["A1"].font = Font(italic=True, color=MUTED.lstrip("#"))
        workbook["Filter State"].freeze_panes = "A3"
        workbook["Filter State"].auto_filter.ref = (
            f"A2:{get_column_letter(len(SAFE_EXPORT_COLUMNS['levoit']))}"
            f"{workbook['Filter State'].max_row}"
        )
    workbook.save(destination)
    # Verification that the resulting file is a real readable workbook.
    verified = load_workbook(destination, read_only=True, data_only=False)
    verified.close()
    return destination


def render_raw_zip(report: PrintExport, destination: str | Path) -> Path:
    destination = Path(destination)
    with ZipFile(destination, "w", compression=ZIP_DEFLATED, compresslevel=6) as archive:
        archive.writestr("print-metadata.csv", _dict_csv([report.print_record]))
        archive.writestr("sgx-samples.csv", _rows_csv(report.sgx_samples, SAFE_EXPORT_COLUMNS["sgx"]))
        archive.writestr("sps30-samples.csv", _rows_csv(report.sps30_samples, SAFE_EXPORT_COLUMNS["sps30"]))
        archive.writestr(
            "summary-metrics.csv",
            _dict_csv([{"metric": key, **metric.as_dict()} for key, metric in report.metrics.items()]),
        )
        archive.writestr("metadata.json", json.dumps(report.metadata(), indent=2, ensure_ascii=False))
        if report.levoit_samples:
            archive.writestr("levoit-samples.csv", _rows_csv(report.levoit_samples, SAFE_EXPORT_COLUMNS["levoit"]))
    return destination


def render_complete_zip(report: PrintExport, destination: str | Path) -> Path:
    destination = Path(destination)
    stem = safe_stem(report)
    with TemporaryDirectory(prefix="airmonitor-package-") as directory:
        root = Path(directory)
        png = render_publication_png(report, root / f"{stem}.png")
        pdf = render_pdf(report, root / f"{stem}.pdf")
        xlsx = render_xlsx(report, root / f"{stem}.xlsx")
        raw = render_raw_zip(report, root / f"{stem}-raw-data.zip")
        metadata = root / "metadata.json"
        metadata.write_text(json.dumps(report.metadata(), indent=2, ensure_ascii=False), encoding="utf-8")
        readme = root / "README.txt"
        readme.write_text(_readme(report), encoding="utf-8")
        with ZipFile(destination, "w", compression=ZIP_DEFLATED, compresslevel=6) as archive:
            for path in (png, pdf, xlsx, raw, metadata, readme):
                archive.write(path, path.name)
    return destination


def export_page(report: PrintExport) -> bytes:
    stem = safe_stem(report)
    buttons = (
        ("Download PDF report", "pdf"),
        ("Download publication PNG", "png"),
        ("Download Excel workbook", "xlsx"),
        ("Download raw CSV ZIP package", "raw"),
        ("Download complete ZIP experiment package", "complete"),
    )
    button_html = "".join(
        f'<a class="download-button" href="/exports/download?print_id={report.print_id}&amp;format={kind}">'
        f"{escape(label)} <span aria-hidden=\"true\">↓</span></a>"
        for label, kind in buttons
    )
    warnings = "".join(f"<li>{escape(item)}</li>" for item in report.warnings)
    dashboard = (
        "/grafana/d/airmonitor-print-window/airmonitor-print-window"
        f"?var-print_id={report.print_id}&amp;theme=light"
    )
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="light"><meta name="theme-color" content="#f4f8fb">
<title>Export {escape(report.title)} - AirMonitor</title>
<link rel="icon" href="/status-assets/favicon.ico">
<link rel="stylesheet" href="/exports/assets/export.css"></head>
<body><main class="export-shell">
<header class="export-hero">
<a href="https://github.com/ddivins/airmonitor" target="_blank" rel="noopener noreferrer">
<img src="/status-assets/airmonitor-logo-300px.webp" alt="AirMonitor"></a>
<div><span class="eyebrow">Print experiment export</span><h1>{escape(report.title)}</h1>
<p>Print #{report.print_id:04d}{' - active when generated' if report.active else ''}</p></div>
</header>
<section class="summary-card"><h2>Print summary</h2>
<dl>
<div><dt>Print start</dt><dd>{escape(_display_time(report.started_at))}</dd></div>
<div><dt>Print end / last seen</dt><dd>{escape(_display_time(report.ended_at))}</dd></div>
<div><dt>Export window</dt><dd>{escape(_format_range(report.window_start, report.window_end))}</dd></div>
<div><dt>Material</dt><dd>{escape(_value(report.print_record.get('filament_type')))}</dd></div>
<div><dt>Profile</dt><dd>{escape(_value(report.print_record.get('filament_profile') or report.print_record.get('filament_sub_brand')))}</dd></div>
<div><dt>Samples</dt><dd>{len(report.sgx_samples):,} SGX / {len(report.sps30_samples):,} SPS30</dd></div>
</dl>{f'<ul class="warnings">{warnings}</ul>' if warnings else ''}</section>
<section><div class="section-heading"><h2>Downloads</h2><span>{escape(stem)}</span></div>
<div class="download-grid">{button_html}</div></section>
<p class="limitation">VOC values are intended for relative and comparative analysis. They are not
compound-specific, regulatory, medical, OSHA-valid, or life-safety exposure measurements.</p>
<nav><a href="{dashboard}">Back to selected Print Window dashboard</a><a href="/">AirMonitor home</a></nav>
</main></body></html>""".encode("utf-8")


def error_page(title: str, message: str, status: int) -> bytes:
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><meta name="color-scheme" content="light">
<title>{escape(title)} - AirMonitor</title><link rel="stylesheet" href="/exports/assets/export.css"></head>
<body><main class="export-shell"><section class="error-card"><span class="eyebrow">Export error {status}</span>
<h1>{escape(title)}</h1><p>{escape(message)}</p><nav><a href="/">AirMonitor home</a>
<a href="/grafana/d/airmonitor-print-window/airmonitor-print-window">Print Window dashboard</a></nav>
</section></main></body></html>""".encode("utf-8")


def _draw_logo(axis: Any) -> None:
    logo = files("airmonitor").joinpath("status_static/airmonitor-logo-300px.webp")
    try:
        with as_file(logo) as logo_path:
            image = mpimg.imread(logo_path)
        inset = axis.inset_axes([0, 0.04, 0.15, 0.92])
        inset.imshow(image)
        inset.axis("off")
    except (OSError, ValueError):
        axis.text(0, 0.56, "AirMonitor", fontsize=24, weight="bold", color=NAVY)
        axis.text(0, 0.24, "Monitor. Understand. Don't Die.", fontsize=8, color=MUTED)


def _plot_series(
    axis: Any,
    report: PrintExport,
    samples: Iterable[dict[str, Any]],
    series: tuple[tuple[str, str, str], ...],
    title: str,
    unit: str,
) -> None:
    axis.set_title(title, loc="left", color=NAVY, weight="bold", fontsize=12)
    plotted = False
    for key, label, color in series:
        points = [(row["sampled_at"], row.get(key)) for row in samples if row.get(key) is not None]
        if points:
            axis.plot([x for x, _ in points], [y for _, y in points], label=label, color=color, linewidth=2)
            plotted = True
    _decorate_time_axis(axis, report, unit)
    if plotted:
        axis.legend(loc="upper left", frameon=False, ncol=min(4, len(series)), fontsize=8)
    else:
        axis.text(0.5, 0.5, "Dataset unavailable", transform=axis.transAxes, ha="center", color=MUTED)


def _plot_environment(axis: Any, report: PrintExport) -> None:
    axis.set_title("Temperature and humidity", loc="left", color=NAVY, weight="bold", fontsize=12)
    temp = [(row["sampled_at"], row.get("temperature_c")) for row in report.sgx_samples if row.get("temperature_c") is not None]
    humid = [(row["sampled_at"], row.get("humidity_rh")) for row in report.sgx_samples if row.get("humidity_rh") is not None]
    if temp:
        axis.plot([x for x, _ in temp], [y for _, y in temp], color=TEAL, linewidth=2, label="Temperature C")
    second = axis.twinx()
    if humid:
        second.plot([x for x, _ in humid], [y for _, y in humid], color=BLUE, linewidth=2, label="Humidity %RH")
    _decorate_time_axis(axis, report, "C")
    second.set_ylabel("%RH", color=MUTED, fontsize=8)
    second.tick_params(labelsize=7, colors=MUTED)
    lines = [
        line for line in axis.get_lines() + second.get_lines()
        if not line.get_label().startswith("_")
    ]
    if lines:
        axis.legend(lines, [line.get_label() for line in lines], loc="upper left", frameon=False, fontsize=7)
    else:
        axis.text(0.5, 0.75, "Dataset unavailable", transform=axis.transAxes, ha="center", color=MUTED)


def _summary_box(axis: Any, report: PrintExport) -> None:
    voc = report.metrics["voc"]
    pm = report.metrics["pm2_5"]
    text = (
        "\n\nSummary\n"
        f"VOC baseline   {_metric(voc.baseline, 'ppm')}\n"
        f"VOC peak       {_metric(voc.peak, 'ppm')}\n"
        f"VOC increase   {_metric(voc.increase, 'ppm')}\n"
        f"VOC time peak  {_duration(voc.time_to_peak_seconds)}\n\n"
        f"PM2.5 baseline {_metric(pm.baseline, 'ug/m3')}\n"
        f"PM2.5 peak     {_metric(pm.peak, 'ug/m3')}\n"
        f"PM2.5 increase {_metric(pm.increase, 'ug/m3')}\n"
        f"PM time peak   {_duration(pm.time_to_peak_seconds)}\n\n"
        f"Samples        {len(report.sgx_samples):,} SGX\n"
        f"               {len(report.sps30_samples):,} SPS30"
    )
    axis.text(
        0.03, 0.03, text, transform=axis.transAxes, va="bottom", color=INK,
        fontsize=8.2, linespacing=1.45, family="monospace",
        bbox={"boxstyle": "round,pad=0.8", "facecolor": "white", "edgecolor": GRID, "alpha": 0.96},
    )


def _decorate_time_axis(axis: Any, report: PrintExport, unit: str) -> None:
    axis.axvspan(report.started_at, report.ended_at, color=BLUE, alpha=0.055)
    axis.axvline(report.started_at, color=NAVY, linestyle="--", linewidth=1.1)
    axis.axvline(report.ended_at, color=NAVY, linestyle=":", linewidth=1.1)
    axis.annotate(
        "Print start", xy=(report.started_at, 1), xycoords=("data", "axes fraction"),
        xytext=(4, -4), textcoords="offset points", color=NAVY, fontsize=6.5,
        ha="left", va="top",
    )
    axis.annotate(
        "Last seen" if report.active else "Print end",
        xy=(report.ended_at, 1), xycoords=("data", "axes fraction"),
        xytext=(-4, -4), textcoords="offset points", color=NAVY, fontsize=6.5,
        ha="right", va="top",
    )
    axis.set_xlim(report.window_start, report.window_end)
    axis.set_ylabel(unit, color=MUTED, fontsize=8)
    axis.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    axis.grid(True, color=GRID, linewidth=0.65, alpha=0.8)
    axis.tick_params(labelsize=7, colors=MUTED)
    for spine in axis.spines.values():
        spine.set_color(GRID)


def _render_pdf_charts(report: PrintExport, destination: Path) -> None:
    fig, axes = plt.subplots(3, 1, figsize=(12, 8), dpi=150, constrained_layout=True)
    fig.patch.set_facecolor("white")
    _plot_series(axes[0], report, report.sgx_samples, (("gas_ppm", "VOC", BLUE),), "VOC", "ppm")
    _plot_series(
        axes[1], report, report.sps30_samples,
        (("mass_pm1_0", "PM1", TEAL), ("mass_pm2_5", "PM2.5", BLUE),
         ("mass_pm4_0", "PM4", "#7754c5"), ("mass_pm10", "PM10", "#e49b00")),
        "Particulate matter mass", "ug/m3",
    )
    temp = [(row["sampled_at"], row.get("temperature_c")) for row in report.sgx_samples if row.get("temperature_c") is not None]
    humid = [(row["sampled_at"], row.get("humidity_rh")) for row in report.sgx_samples if row.get("humidity_rh") is not None]
    if temp:
        axes[2].plot([x for x, _ in temp], [y for _, y in temp], label="Temperature C", color=TEAL)
    if humid:
        axes[2].plot([x for x, _ in humid], [y for _, y in humid], label="Humidity %RH", color=BLUE)
    axes[2].set_title("Temperature and humidity", loc="left", color=NAVY, weight="bold")
    _decorate_time_axis(axes[2], report, "C / %RH")
    if temp or humid:
        axes[2].legend(frameon=False)
    fig.savefig(destination, dpi=150, facecolor="white")
    plt.close(fig)


def _pdf_table(rows: list[tuple[str, str]], widths: tuple[float, float]) -> Table:
    table = Table(rows, colWidths=list(widths), hAlign="LEFT")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(NAVY)),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor(INK)),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor(GRID)),
    ]))
    return table


def _metric_table(report: PrintExport) -> Table:
    rows = [["Metric", "Pre-print median", "Peak during print", "Increase", "Time to peak", "Samples"]]
    for metric in report.metrics.values():
        rows.append([
            metric.name,
            _metric(metric.baseline, metric.unit),
            _metric(metric.peak, metric.unit),
            _metric(metric.increase, metric.unit),
            _duration(metric.time_to_peak_seconds),
            f"{metric.total_samples:,}",
        ])
    table = Table(rows, colWidths=[1.35 * inch, 1.55 * inch, 1.55 * inch, 1.35 * inch, 1.25 * inch, 0.9 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (0, -1), colors.HexColor(NAVY)),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(GRID)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f8fb")]),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return table


def _pdf_page(canvas: Any, doc: Any) -> None:
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor(GRID))
    canvas.line(0.55 * inch, 0.34 * inch, landscape(letter)[0] - 0.55 * inch, 0.34 * inch)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor(MUTED))
    canvas.drawString(0.55 * inch, 0.19 * inch, "AirMonitor - Monitor. Understand. Don't Die.")
    canvas.drawRightString(landscape(letter)[0] - 0.55 * inch, 0.19 * inch, f"Page {doc.page}")
    canvas.restoreState()


def _summary_sheet(sheet: Any, report: PrintExport) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "AirMonitor Print Experiment Export"
    sheet["A1"].font = Font(size=20, bold=True, color=NAVY.lstrip("#"))
    sheet.merge_cells("A1:D1")
    rows = [
        ("Print ID", report.print_id, "", ""),
        ("Print title", report.title, "", ""),
        ("Active when generated", report.active, "", ""),
        ("Print start (UTC)", _excel_time(report.started_at), "", ""),
        ("Print end / last seen (UTC)", _excel_time(report.ended_at), "", ""),
        ("Export window start (UTC)", _excel_time(report.window_start), "", ""),
        ("Export window end (UTC)", _excel_time(report.window_end), "", ""),
        ("Material", report.print_record.get("filament_type"), "", ""),
        ("Filament profile", report.print_record.get("filament_profile") or report.print_record.get("filament_sub_brand"), "", ""),
        ("Filament color", report.print_record.get("filament_color"), "", ""),
        ("", "", "", ""),
        ("Metric", "Pre-print median", "Peak during print", "Increase / time to peak"),
    ]
    for metric in report.metrics.values():
        rows.append((
            f"{metric.name} ({metric.unit})",
            metric.baseline,
            metric.peak,
            f"{_number(metric.increase)} / {_duration(metric.time_to_peak_seconds)}",
        ))
    rows.extend([
        ("", "", "", ""),
        ("SGX sample count", len(report.sgx_samples), "", ""),
        ("SPS30 sample count", len(report.sps30_samples), "", ""),
        ("Levoit sample count", len(report.levoit_samples), "", ""),
        ("Baseline methodology", "Median of valid samples in the 30-minute pre-print portion", "", ""),
        ("Window methodology", "30 minutes before print start through 30 minutes after end/last seen", "", ""),
        ("Interpretation", "VOC is relative/comparative and not compound-specific or regulatory.", "", ""),
    ])
    for row in rows:
        sheet.append(row)
    for cell in sheet[14]:
        cell.fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
        cell.font = Font(bold=True, color="FFFFFF")
    for row in sheet.iter_rows(min_row=2, max_col=1):
        row[0].font = Font(bold=True, color=NAVY.lstrip("#"))
    for row in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "")
        for column in (2, 3):
            value = sheet.cell(row, column).value
            if isinstance(value, (float, int)) and not isinstance(value, bool):
                sheet.cell(row, column).number_format = (
                    "0" if label == "Print ID" or "sample count" in label.lower() else "0.000"
                )
        if isinstance(sheet.cell(row, 2).value, datetime):
            sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 31
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 28
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _page_setup(sheet)


def _raw_sheet(sheet: Any, rows: Iterable[dict[str, Any]], columns: tuple[str, ...]) -> None:
    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
        cell.font = Font(bold=True, color="FFFFFF")
    for row in rows:
        sheet.append([_excel_value(row.get(column)) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, sheet.max_row)}"
    for index, column in enumerate(columns, 1):
        is_timestamp = column == "sampled_at" or column.endswith("_at")
        width = 22 if is_timestamp else min(28, max(12, len(column) + 3))
        sheet.column_dimensions[get_column_letter(index)].width = width
        if is_timestamp:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value in cell:
                    value.number_format = "yyyy-mm-dd hh:mm:ss.000"
        elif any(token in column for token in ("ppm", "mass_", "number_", "temperature", "humidity", "particle_size", "pm2_5")):
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value in cell:
                    value.number_format = "0.000"
    _page_setup(sheet)


def _metadata_sheet(sheet: Any, report: PrintExport) -> None:
    sheet.append(("Field", "Value"))
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
        cell.font = Font(bold=True, color="FFFFFF")
    for key, value in report.print_record.items():
        sheet.append((key, _excel_value(value)))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 70
    _page_setup(sheet)


def _page_setup(sheet: Any) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4


def _rows_csv(rows: Iterable[dict[str, Any]], columns: tuple[str, ...]) -> str:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _csv_value(row.get(column)) for column in columns})
    return output.getvalue()


def _dict_csv(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return _rows_csv(rows, tuple(columns))


def _readme(report: PrintExport) -> str:
    return f"""AirMonitor complete print experiment package

Print: {report.title}
Print ID: {report.print_id}
Print start: {report.started_at.isoformat()}
Print end / last seen: {report.ended_at.isoformat()}
Export window: {report.window_start.isoformat()} through {report.window_end.isoformat()}

Files:
- PNG: high-resolution publication graphic
- PDF: multipage experiment report
- XLSX: summary and timestamped data worksheets
- raw-data ZIP: CSV datasets and machine-readable metadata
- metadata.json: calculations, provenance, warnings, and print metadata

Units:
- VOC: ppm reported by the cross-sensitive SGX TVOC sensor
- Temperature: degrees Celsius
- Humidity: percent relative humidity
- Particulate mass: micrograms per cubic meter
- Particle counts: particles per cubic centimeter
- Typical particle size: micrometers

Baselines are medians of valid samples collected during the 30-minute pre-print period.
The export window extends 30 minutes before print start through 30 minutes after print
end or last-seen time.

Measurement limitations:
VOC values are intended for relative and comparative analysis. They are not
compound-specific, regulatory, medical, OSHA-valid, or life-safety exposure
measurements. AirMonitor is a DIY monitoring platform, not a certified instrument.
"""


def _format_range(start: datetime, end: datetime) -> str:
    return f"{_display_time(start)} to {_display_time(end)}"


def _display_time(value: datetime) -> str:
    return value.astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def _material_line(report: PrintExport) -> str:
    values = [
        report.print_record.get("filament_type"),
        report.print_record.get("filament_profile") or report.print_record.get("filament_sub_brand"),
        report.print_record.get("filament_color"),
    ]
    return " | ".join(str(value) for value in values if value) or "Filament metadata not captured"


def _value(value: Any) -> str:
    return str(value) if value not in (None, "") else "Not captured"


def _number(value: float | None) -> str:
    return "Unavailable" if value is None or not math.isfinite(value) else f"{value:.3f}"


def _metric(value: float | None, unit: str) -> str:
    return "Unavailable" if value is None or not math.isfinite(value) else f"{value:.3f} {unit}"


def _duration(value: float | None) -> str:
    if value is None:
        return "Unavailable"
    minutes, seconds = divmod(max(0, int(round(value))), 60)
    return f"{minutes}m {seconds:02d}s"


def _excel_time(value: datetime) -> datetime:
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _excel_value(value: Any) -> Any:
    return _excel_time(value) if isinstance(value, datetime) else value


def _csv_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat().replace("+00:00", "Z")
    if isinstance(value, bool):
        return "true" if value else "false"
    return value
