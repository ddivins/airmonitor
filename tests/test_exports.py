from __future__ import annotations

from datetime import datetime, timezone
import io
import json
from pathlib import Path
import threading
from zipfile import ZipFile

from openpyxl import load_workbook
from PIL import Image

from airmonitor.database import connect
from airmonitor.exports.renderers import (
    export_page,
    render_complete_zip,
    render_pdf,
    render_publication_png,
    render_raw_zip,
    render_xlsx,
    safe_stem,
)
from airmonitor.exports.repository import ExportNotFound, ExportRepository
from airmonitor.exports.web import ExportHandler
from airmonitor.sps30_service import ensure_schema


NOW = datetime(2026, 7, 16, 16, 0, tzinfo=timezone.utc)


def fixture_database(path: Path, *, active: bool = False, sgx: bool = True, sps: bool = True) -> int:
    conn = connect(path)
    ensure_schema(conn)
    conn.execute(
        """
        INSERT INTO prints (
            id, started_at, ended_at, last_seen_at, started_gcode_state,
            last_gcode_state, ended_gcode_state, subtask_name, filament_type,
            filament_color, filament_profile, filament_sub_brand
        ) VALUES (26, '2026-07-16T14:00:00.000Z', ?, '2026-07-16T15:00:00.000Z',
                  'RUNNING', ?, ?, '../Example / Blades', 'ABS', 'FF0000',
                  'Generic ABS', 'AirMonitor Test')
        """,
        (
            None if active else "2026-07-16T15:00:00.000Z",
            "RUNNING" if active else "FINISH",
            None if active else "FINISH",
        ),
    )
    conn.execute(
        """
        INSERT INTO sensors(sensor_id, manufacturer, product, model, transport)
        VALUES ('sgx-01','Amphenol','PS1','PS1','usb'),
               ('sps30-01','Sensirion','SPS30','SPS30','usb')
        """
    )
    if sgx:
        rows = [
            ("2026-07-16T13:29:59.000Z", 99.0),
            ("2026-07-16T13:30:00.000Z", 0.10),
            ("2026-07-16T13:45:00.000Z", 0.20),
            ("2026-07-16T13:59:59.000Z", 0.30),
            ("2026-07-16T14:10:00.000Z", 0.70),
            ("2026-07-16T14:20:00.000Z", 1.20),
            ("2026-07-16T15:20:00.000Z", 0.40),
            ("2026-07-16T15:30:00.000Z", 0.25),
            ("2026-07-16T15:30:01.000Z", 88.0),
        ]
        conn.executemany(
            """
            INSERT INTO sgx_voc_samples (
                sampled_at, sensor_id, sensor_protocol, sensor_port, gas_ppm,
                gas_mass, full_scale, temperature_c, humidity_rh
            ) VALUES (?, 'sgx-01', 'uart', '/dev/test-sgx', ?, 0.0, 1000, 23.5, 45.0)
            """,
            rows,
        )
    if sps:
        rows = [
            ("2026-07-16T13:30:00.000Z", 1.0),
            ("2026-07-16T13:45:00.000Z", 2.0),
            ("2026-07-16T14:05:00.000Z", 5.0),
            ("2026-07-16T14:25:00.000Z", 9.0),
            ("2026-07-16T15:30:00.000Z", 3.0),
        ]
        conn.executemany(
            """
            INSERT INTO sps30_samples (
                sampled_at, sensor_id, sensor_port, mass_pm1_0, mass_pm2_5,
                mass_pm4_0, mass_pm10, number_pm0_5, number_pm1_0,
                number_pm2_5, number_pm4_0, number_pm10, typical_particle_size
            ) VALUES (?, 'sps30-01', '/dev/test-sps', ?, ?, ?, ?, 10, 9, 8, 7, 6, 0.5)
            """,
            [(when, value * 0.8, value, value * 1.1, value * 1.2) for when, value in rows],
        )
    conn.execute(
        """
        INSERT INTO levoit_samples (
            sampled_at, device_name, power_state, mode, fan_level, pm2_5,
            air_quality, filter_life_percent
        ) VALUES ('2026-07-16T14:30:00.000Z','Core 400S','on','manual',4,7,1,92)
        """
    )
    conn.execute(
        """
        INSERT INTO filter_control_state (
            filter_id, manual_mode, automation_request, actual_state,
            effective_state, reason
        ) VALUES ('levoit','auto','on','on','on','test fixture')
        """
    )
    conn.commit()
    conn.close()
    return 26


def test_completed_print_uses_exact_window_and_median_baseline(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database)
    report = ExportRepository(str(database)).load(26, now=NOW)
    assert report.window_start.isoformat() == "2026-07-16T13:30:00+00:00"
    assert report.window_end.isoformat() == "2026-07-16T15:30:00+00:00"
    assert [row["gas_ppm"] for row in report.sgx_samples] == [0.1, 0.2, 0.3, 0.7, 1.2, 0.4, 0.25]
    assert report.metrics["voc"].baseline == 0.2
    assert report.metrics["voc"].peak == 1.2
    assert report.metrics["voc"].increase == 1.0
    assert report.metrics["voc"].time_to_peak_seconds == 1200
    assert report.metrics["pm2_5"].baseline == 1.5
    assert report.metrics["pm2_5"].peak == 9.0


def test_active_print_is_labeled_and_uses_last_seen(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database, active=True)
    report = ExportRepository(str(database)).load(26, now=NOW)
    assert report.active
    assert report.ended_at.isoformat() == "2026-07-16T15:00:00+00:00"
    assert "preliminary" in " ".join(report.warnings)


def test_missing_print_and_missing_sensor_streams(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database, sgx=False, sps=False)
    repository = ExportRepository(str(database))
    report = repository.load(26, now=NOW)
    assert not report.sgx_samples
    assert not report.sps30_samples
    assert report.metrics["voc"].baseline is None
    assert len(report.warnings) == 2
    try:
        repository.load(999, now=NOW)
    except ExportNotFound:
        pass
    else:
        raise AssertionError("missing print should fail")


def test_safe_filename_and_public_page(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database)
    report = ExportRepository(str(database)).load(26, now=NOW)
    assert safe_stem(report) == "airmonitor-print-0026-example-blades-2026-07-16"
    page = export_page(report).decode()
    assert "/exports/download?print_id=26&amp;format=pdf" in page
    assert "var-print_id=26" in page
    assert "../" not in safe_stem(report)


def test_public_export_http_routes_and_concurrency_limit(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database)
    server = type("Server", (), {
        "repository": ExportRepository(str(database)),
        "generation_lock": threading.BoundedSemaphore(1),
    })()
    page = handler_request(server, "/print?print_id=26")
    assert page.startswith(b"HTTP/1.0 200")
    assert b"Download PDF report" in page
    raw = handler_request(server, "/download?print_id=26&format=raw")
    assert raw.startswith(b"HTTP/1.0 200")
    assert b"Content-Type: application/zip" in raw
    assert b"airmonitor-print-0026" in raw
    server.generation_lock.acquire()
    try:
        busy = handler_request(server, "/download?print_id=26&format=pdf")
        assert busy.startswith(b"HTTP/1.0 429")
        assert b"Retry-After: 10" in busy
    finally:
        server.generation_lock.release()
    invalid = handler_request(server, "/print?print_id=not-a-number")
    assert invalid.startswith(b"HTTP/1.0 400")


class FakeSocket:
    def __init__(self, request: bytes) -> None:
        self.request = io.BytesIO(request)
        self.response = io.BytesIO()

    def makefile(self, mode: str, _buffering: int | None = None):
        return self.request if "r" in mode else self.response

    def sendall(self, data: bytes) -> None:
        self.response.write(data)


def handler_request(server, path: str) -> bytes:
    socket = FakeSocket(f"GET {path} HTTP/1.0\r\nHost: test\r\n\r\n".encode())
    ExportHandler(socket, ("127.0.0.1", 12345), server)
    return socket.response.getvalue()


def test_png_pdf_xlsx_and_zip_artifacts(tmp_path: Path) -> None:
    database = tmp_path / "airmonitor.sqlite3"
    fixture_database(database)
    report = ExportRepository(str(database)).load(26, now=NOW)
    png = render_publication_png(report, tmp_path / "report.png")
    pdf = render_pdf(report, tmp_path / "report.pdf")
    xlsx = render_xlsx(report, tmp_path / "report.xlsx")
    raw = render_raw_zip(report, tmp_path / "raw.zip")
    complete = render_complete_zip(report, tmp_path / "complete.zip")

    with Image.open(png) as image:
        assert image.format == "PNG"
        assert image.size == (3840, 2160)
    assert pdf.read_bytes().startswith(b"%PDF")
    assert pdf.stat().st_size > 20_000
    workbook = load_workbook(xlsx, read_only=True, data_only=False)
    assert workbook.sheetnames == ["Summary", "SGX Samples", "SPS30 Samples", "Print Metadata", "Filter State"]
    assert workbook["Summary"]["B2"].value == 26
    assert isinstance(workbook["SGX Samples"]["B2"].value, datetime)
    assert workbook["SGX Samples"]["K2"].value == 23.5
    assert workbook["SGX Samples"]["K2"].number_format == "0.000"
    workbook.close()
    with ZipFile(raw) as archive:
        assert {
            "print-metadata.csv", "sgx-samples.csv", "sps30-samples.csv",
            "summary-metrics.csv", "metadata.json", "levoit-samples.csv",
        } <= set(archive.namelist())
        metadata = json.loads(archive.read("metadata.json"))
        assert metadata["print_id"] == 26
    with ZipFile(complete) as archive:
        names = set(archive.namelist())
        assert "README.txt" in names
        assert "metadata.json" in names
        assert any(name.endswith(".png") for name in names)
        assert any(name.endswith(".pdf") for name in names)
        assert any(name.endswith(".xlsx") for name in names)
        assert any(name.endswith("-raw-data.zip") for name in names)
